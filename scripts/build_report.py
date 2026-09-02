#!/usr/bin/env python3

"""
FIRIS - Detailed Excel Fire Risk Report
=======================================

این فایل شاخص FLI را محاسبه نمی‌کند.

فقط Raster نهایی FLI را می‌خواند و گزارش Excel می‌سازد.

خروجی:

1. خلاصه
2. مناطق_چهارگانه
3. مناطق_شکار_ممنوع
4. طبقات_خطر
5. اطلاعات_فنی

نگاشت قطعی لایه‌ها:

protected_areas.geojson
    -> مناطق چهارگانه

hunting_banned.geojson
    -> مناطق شکار ممنوع
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import numpy as np
import rasterio

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from rasterio.mask import mask


# ============================================================
# RISK CLASSES
# ============================================================

RISK_CLASSES = [
    ("کم", 0.0, 20.0),
    ("متوسط", 20.0, 40.0),
    ("زیاد", 40.0, 60.0),
    ("خیلی زیاد", 60.0, 80.0),
    ("بحرانی", 80.0, 100.000001),
]


RISK_COLORS = {
    "کم": "2E7D32",
    "متوسط": "FDD835",
    "زیاد": "FB8C00",
    "خیلی زیاد": "E53935",
    "بحرانی": "880E4F",
    "بدون داده": "777777",
}


# ============================================================
# ARGUMENTS
# ============================================================

def parse_args():

    parser = argparse.ArgumentParser(
        description="Build detailed FIRIS fire-risk Excel report."
    )

    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="Existing FLI raster."
    )

    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Output Excel file."
    )

    parser.add_argument(
        "--run-date",
        required=False,
        default=None,
        help="Forecast date YYYY-MM-DD."
    )

    parser.add_argument(
        "--protected",
        required=False,
        type=Path,
        default=Path("protected_areas.geojson"),
        help="Four protected areas GeoJSON."
    )

    parser.add_argument(
        "--hunting",
        required=False,
        type=Path,
        default=Path("hunting_banned.geojson"),
        help="Hunting banned areas GeoJSON."
    )

    return parser.parse_args()


# ============================================================
# FILE CHECK
# ============================================================

def require_file(
    path: Path,
    label: str
):

    if not path.is_file():

        raise FileNotFoundError(
            f"{label} not found: {path}"
        )


# ============================================================
# RISK CLASS
# ============================================================

def risk_class(
    value: float | None
) -> str:

    if value is None:
        return "بدون داده"

    value = float(value)

    if not np.isfinite(value):
        return "بدون داده"

    if value < 20:
        return "کم"

    if value < 40:
        return "متوسط"

    if value < 60:
        return "زیاد"

    if value < 80:
        return "خیلی زیاد"

    return "بحرانی"


# ============================================================
# RISK FILL
# ============================================================

def risk_fill(
    label: str
) -> PatternFill:

    return PatternFill(
        fill_type="solid",
        fgColor=RISK_COLORS.get(
            label,
            RISK_COLORS["بدون داده"]
        )
    )


# ============================================================
# RISK FONT
# ============================================================

def risk_font(
    label: str
) -> Font:

    # زرد باید نوشته مشکی داشته باشد
    if label == "متوسط":
        text_color = "000000"
    else:
        text_color = "FFFFFF"

    return Font(
        name="B Nazanin",
        size=11,
        bold=True,
        color=text_color
    )


# ============================================================
# APPLY RISK STYLE
# ============================================================

def apply_risk_style(
    cell,
    label: str
):

    cell.fill = risk_fill(label)
    cell.font = risk_font(label)

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )


# ============================================================
# LOAD GEOJSON
# ============================================================

def load_geojson(
    path: Path
) -> dict[str, Any]:

    require_file(
        path,
        "GeoJSON"
    )

    with path.open(
        "r",
        encoding="utf-8"
    ) as file:

        data = json.load(file)

    if data.get("type") != "FeatureCollection":

        raise ValueError(
            f"Invalid GeoJSON FeatureCollection: {path}"
        )

    features = data.get(
        "features",
        []
    )

    if not features:

        raise ValueError(
            f"No features found in: {path}"
        )

    return data


# ============================================================
# GEOJSON FINGERPRINT
# ============================================================

def geojson_fingerprint(
    data: dict[str, Any]
) -> str:

    """
    برای تشخیص اینکه دو فایل GeoJSON
    عملاً محتوای یکسان دارند یا نه.
    """

    normalized = json.dumps(
        data,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":")
    )

    return normalized


# ============================================================
# CHECK GEOJSON DUPLICATION
# ============================================================

def check_geojson_sources(
    protected_geojson: dict[str, Any],
    hunting_geojson: dict[str, Any],
    protected_path: Path,
    hunting_path: Path
):

    protected_fp = geojson_fingerprint(
        protected_geojson
    )

    hunting_fp = geojson_fingerprint(
        hunting_geojson
    )

    if protected_fp == hunting_fp:

        print("")
        print("=" * 70)
        print("WARNING: DUPLICATE GEOJSON DATA")
        print("=" * 70)
        print("")
        print(
            "protected_areas.geojson and hunting_banned.geojson "
            "contain identical data."
        )
        print("")
        print(
            "Excel mapping is correct, but the source data are identical."
        )
        print("")
        print(
            f"Protected: {protected_path}"
        )
        print(
            f"Hunting:   {hunting_path}"
        )
        print("")
        print(
            "The report will still be generated, but both sheets "
            "will naturally contain the same geographic features."
        )
        print("=" * 70)
        print("")


# ============================================================
# FEATURE NAME
# ============================================================

def get_feature_name(
    feature: dict[str, Any],
    fallback: str
) -> str:

    properties = (
        feature.get("properties")
        or {}
    )

    candidates = [
        "name",
        "NAME",
        "Name",
        "shapeName",
        "ShapeName",
        "NAME_1",
        "NAME_FA",
        "name_fa",
        "title",
        "TITLE",
        "نام",
        "نام منطقه",
        "نام_منطقه",
    ]

    for key in candidates:

        value = properties.get(key)

        if (
            value is not None
            and str(value).strip()
        ):

            return str(
                value
            ).strip()

    for key, value in properties.items():

        if not isinstance(
            value,
            str
        ):
            continue

        if not value.strip():
            continue

        key_lower = (
            str(key)
            .strip()
            .lower()
        )

        if (
            "name" in key_lower
            or "title" in key_lower
            or "نام" in key_lower
        ):

            return value.strip()

    return fallback


# ============================================================
# READ FLI
# ============================================================

def load_fli(
    path: Path
):

    with rasterio.open(path) as src:

        if src.crs is None:

            raise ValueError(
                "FLI raster has no CRS."
            )

        data = src.read(
            1
        ).astype(
            np.float32
        )

        if src.nodata is not None:

            data[
                np.isclose(
                    data,
                    float(src.nodata)
                )
            ] = np.nan

        data[
            ~np.isfinite(data)
        ] = np.nan

        data[
            (data < 0)
            |
            (data > 100)
        ] = np.nan

        valid = np.isfinite(data)

        if not np.any(valid):

            raise RuntimeError(
                "FLI raster contains no valid pixels."
            )

        valid_values = data[valid]

        statistics = {
            "min": float(
                np.min(valid_values)
            ),
            "max": float(
                np.max(valid_values)
            ),
            "mean": float(
                np.mean(valid_values)
            ),
            "count": int(
                valid_values.size
            ),
        }

        metadata = {
            "crs": str(src.crs),
            "width": int(src.width),
            "height": int(src.height),
            "resolution_x": float(
                src.res[0]
            ),
            "resolution_y": float(
                abs(src.res[1])
            ),
            "bounds": {
                "left": float(
                    src.bounds.left
                ),
                "bottom": float(
                    src.bounds.bottom
                ),
                "right": float(
                    src.bounds.right
                ),
                "top": float(
                    src.bounds.top
                ),
            },
            "transform": [
                float(src.transform.a),
                float(src.transform.b),
                float(src.transform.c),
                float(src.transform.d),
                float(src.transform.e),
                float(src.transform.f),
            ],
        }

    return (
        data,
        statistics,
        metadata
    )


# ============================================================
# REGION STATISTICS
# ============================================================

def calculate_region_statistics(
    src,
    feature: dict[str, Any]
):

    geometry = feature.get(
        "geometry"
    )

    if not geometry:

        return {
            "count": 0,
            "min": None,
            "max": None,
            "mean": None,
            "risk": "بدون داده",
        }

    try:

        masked_data, _ = mask(
            src,
            [geometry],
            crop=True,
            filled=False,
            all_touched=False,
        )

        band = masked_data[0]

        if np.ma.isMaskedArray(band):

            values = (
                band
                .compressed()
                .astype(np.float32)
            )

        else:

            values = np.asarray(
                band,
                dtype=np.float32
            ).ravel()

        values = values[
            np.isfinite(values)
        ]

        values = values[
            (values >= 0)
            &
            (values <= 100)
        ]

        if values.size == 0:

            return {
                "count": 0,
                "min": None,
                "max": None,
                "mean": None,
                "risk": "بدون داده",
            }

        mean_value = float(
            np.mean(values)
        )

        return {
            "count": int(
                values.size
            ),
            "min": float(
                np.min(values)
            ),
            "max": float(
                np.max(values)
            ),
            "mean": mean_value,
            "risk": risk_class(
                mean_value
            ),
        }

    except Exception as error:

        print(
            "WARNING: Region statistics failed:"
        )

        print(
            f"  {error}"
        )

        return {
            "count": 0,
            "min": None,
            "max": None,
            "mean": None,
            "risk": "بدون داده",
        }


# ============================================================
# WHOLE PROVINCE RISK CLASSES
# ============================================================

def calculate_class_statistics(
    values: np.ndarray
):

    valid = (
        np.isfinite(values)
        &
        (values >= 0)
        &
        (values <= 100)
    )

    total = int(
        np.count_nonzero(valid)
    )

    rows = []

    for (
        label,
        minimum,
        maximum
    ) in RISK_CLASSES:

        class_mask = (
            valid
            &
            (values >= minimum)
            &
            (values < maximum)
        )

        count = int(
            np.count_nonzero(
                class_mask
            )
        )

        percent = (
            100.0
            * count
            / total
            if total > 0
            else 0.0
        )

        rows.append(
            {
                "label": label,
                "min": minimum,
                "max": min(
                    maximum,
                    100.0
                ),
                "count": count,
                "percent": percent,
            }
        )

    return rows


# ============================================================
# EXCEL STYLES
# ============================================================

def create_styles():

    thin = Side(
        style="thin",
        color="C9CED3"
    )

    return {
        "border": Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        ),

        "header_fill": PatternFill(
            fill_type="solid",
            fgColor="8B0000",
        ),

        "title_font": Font(
            name="B Nazanin",
            size=16,
            bold=True,
        ),

        "header_font": Font(
            name="B Nazanin",
            size=11,
            bold=True,
            color="FFFFFF",
        ),

        "normal_font": Font(
            name="B Nazanin",
            size=11,
        ),

        "bold_font": Font(
            name="B Nazanin",
            size=11,
            bold=True,
        ),
    }


# ============================================================
# FORMAT SHEET
# ============================================================

def apply_sheet_format(
    worksheet,
    styles
):

    worksheet.sheet_view.rightToLeft = True

    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.page_setup.fitToWidth = 1

    worksheet.page_setup.fitToHeight = 0

    for row in worksheet.iter_rows():

        for cell in row:

            if isinstance(
                cell,
                MergedCell
            ):
                continue

            if cell.value is None:
                continue

            cell.border = styles[
                "border"
            ]

            cell.alignment = Alignment(
                horizontal="right",
                vertical="center",
                wrap_text=True,
            )

            cell.font = styles[
                "normal_font"
            ]


# ============================================================
# HEADER STYLE
# ============================================================

def style_header_row(
    worksheet,
    row_number,
    styles
):

    for cell in worksheet[row_number]:

        if isinstance(
            cell,
            MergedCell
        ):
            continue

        if cell.value is None:
            continue

        cell.fill = styles[
            "header_fill"
        ]

        cell.font = styles[
            "header_font"
        ]

        cell.border = styles[
            "border"
        ]

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


# ============================================================
# AUTO FIT
# ============================================================

def auto_fit_columns(
    worksheet,
    minimum=12,
    maximum=42
):

    column_widths = {}

    for row in worksheet.iter_rows():

        for cell in row:

            if isinstance(
                cell,
                MergedCell
            ):
                continue

            if cell.value is None:
                continue

            try:

                column_letter = (
                    get_column_letter(
                        cell.column
                    )
                )

            except Exception:

                continue

            value_length = len(
                str(cell.value)
            )

            current = column_widths.get(
                column_letter,
                minimum
            )

            column_widths[
                column_letter
            ] = min(
                max(
                    current,
                    value_length + 2
                ),
                maximum
            )

    for column_letter, width in column_widths.items():

        worksheet.column_dimensions[
            column_letter
        ].width = max(
            minimum,
            width
        )


# ============================================================
# ADD TITLE
# ============================================================

def add_sheet_title(
    worksheet,
    title,
    styles,
    columns=6
):

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=columns
    )

    cell = worksheet.cell(
        row=1,
        column=1
    )

    cell.value = title

    cell.font = styles[
        "title_font"
    ]

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[
        1
    ].height = 28


# ============================================================
# SUMMARY SHEET
# ============================================================

def build_summary_sheet(
    workbook,
    fli_stats,
    metadata,
    run_date,
    styles
):

    worksheet = workbook.create_sheet(
        "خلاصه"
    )

    add_sheet_title(
        worksheet,
        "گزارش شاخص خطر حریق FIRIS",
        styles,
        4
    )

    rows = [
        ("عنوان", "شاخص خطر حریق FIRIS"),
        ("تاریخ پیش‌بینی", run_date or "نامشخص"),
        ("حداقل FLI", fli_stats["min"]),
        ("حداکثر FLI", fli_stats["max"]),
        ("میانگین FLI", fli_stats["mean"]),
        ("تعداد سلول‌های معتبر", fli_stats["count"]),
        ("سیستم مختصات", metadata["crs"]),
        (
            "ابعاد Raster",
            f'{metadata["width"]} × {metadata["height"]}'
        ),
        (
            "تفکیک مکانی",
            f'{metadata["resolution_x"]} × '
            f'{metadata["resolution_y"]}'
        ),
    ]

    worksheet.append([])

    worksheet.append([
        "عنوان",
        "مقدار"
    ])

    style_header_row(
        worksheet,
        3,
        styles
    )

    for title, value in rows:

        worksheet.append([
            title,
            value
        ])

    # --------------------------------------------------------
    # قالب‌بندی عمومی
    # --------------------------------------------------------

    apply_sheet_format(
        worksheet,
        styles
    )

    # --------------------------------------------------------
    # عنوان ستون‌ها را دوباره حفظ می‌کنیم
    # --------------------------------------------------------

    style_header_row(
        worksheet,
        3,
        styles
    )

    # --------------------------------------------------------
    # تشخیص طبقه خطر استانی
    # --------------------------------------------------------

    province_risk = risk_class(
        fli_stats["mean"]
    )

    # --------------------------------------------------------
    # ردیف میانگین FLI
    #
    # در جدول فعلی:
    #
    # row 8 = میانگین FLI
    #
    # چون عنوان در ستون A و مقدار در ستون B است.
    # --------------------------------------------------------

    mean_row = None

    for row in range(
        4,
        worksheet.max_row + 1
    ):

        if worksheet.cell(
            row=row,
            column=1
        ).value == "میانگین FLI":

            mean_row = row
            break

    # --------------------------------------------------------
    # رنگ مقدار میانگین FLI
    # --------------------------------------------------------

    if mean_row is not None:

        risk_cell = worksheet.cell(
            row=mean_row,
            column=2
        )

        apply_risk_style(
            risk_cell,
            province_risk
        )

    # --------------------------------------------------------
    # یک ردیف جداگانه برای طبقه خطر استان
    # --------------------------------------------------------

    risk_row = worksheet.max_row + 1

    worksheet.cell(
        row=risk_row,
        column=1
    ).value = "طبقه خطر استان"

    worksheet.cell(
        row=risk_row,
        column=2
    ).value = province_risk

    worksheet.cell(
        row=risk_row,
        column=1
    ).font = styles[
        "bold_font"
    ]

    worksheet.cell(
        row=risk_row,
        column=1
    ).alignment = Alignment(
        horizontal="right",
        vertical="center"
    )

    # رنگ قطعی سلول طبقه خطر
    apply_risk_style(
        worksheet.cell(
            row=risk_row,
            column=2
        ),
        province_risk
    )

    # --------------------------------------------------------
    # دوباره border و alignment سلول‌های جدول
    # --------------------------------------------------------

    for row in worksheet.iter_rows():

        for cell in row:

            if isinstance(
                cell,
                MergedCell
            ):
                continue

            if cell.value is None:
                continue

            cell.border = styles[
                "border"
            ]

    # --------------------------------------------------------
    # دوباره رنگ سلول خطر
    #
    # این قسمت عمداً در انتهای تابع قرار گرفته تا هیچ
    # قالب‌بندی عمومی دیگری رنگ را حذف نکند.
    # --------------------------------------------------------

    if mean_row is not None:

        apply_risk_style(
            worksheet.cell(
                row=mean_row,
                column=2
            ),
            province_risk
        )

        worksheet.cell(
            row=mean_row,
            column=2
        ).border = styles[
            "border"
        ]

    apply_risk_style(
        worksheet.cell(
            row=risk_row,
            column=2
        ),
        province_risk
    )

    worksheet.cell(
        row=risk_row,
        column=2
    ).border = styles[
        "border"
    ]

    auto_fit_columns(
        worksheet,
        minimum=15,
        maximum=42
    )


# ============================================================
# REGION SHEET
# ============================================================

def build_region_sheet(
    workbook,
    sheet_name: str,
    region_type: str,
    geojson: dict[str, Any],
    src,
    styles
):

    worksheet = workbook.create_sheet(
        sheet_name
    )

    add_sheet_title(
        worksheet,
        region_type,
        styles,
        7
    )

    worksheet.append([])

    # --------------------------------------------------------
    # ستون‌ها
    #
    # این ساختار برای هر دو شیت کاملاً یکسان است.
    # فقط region_type تغییر می‌کند.
    # --------------------------------------------------------

    headers = [
        "ردیف",
        "نوع منطقه",
        "نام منطقه",
        "حداقل FLI",
        "حداکثر FLI",
        "میانگین FLI",
        "طبقه خطر",
    ]

    worksheet.append(
        headers
    )

    style_header_row(
        worksheet,
        3,
        styles
    )

    features = geojson.get(
        "features",
        []
    )

    for index, feature in enumerate(
        features,
        start=1
    ):

        name = get_feature_name(
            feature,
            f"منطقه {index}"
        )

        stats = calculate_region_statistics(
            src,
            feature
        )

        row = worksheet.max_row + 1

        worksheet.append([
            index,
            region_type,
            name,
            stats["min"],
            stats["max"],
            stats["mean"],
            stats["risk"],
        ])

        risk_cell = worksheet.cell(
            row=row,
            column=7
        )

        apply_risk_style(
            risk_cell,
            stats["risk"]
        )

    # --------------------------------------------------------
    # قالب عمومی
    # --------------------------------------------------------

    apply_sheet_format(
        worksheet,
        styles
    )

    # --------------------------------------------------------
    # Header را دوباره اعمال می‌کنیم
    # --------------------------------------------------------

    style_header_row(
        worksheet,
        3,
        styles
    )

    # --------------------------------------------------------
    # رنگ طبقه خطر را بعد از قالب عمومی دوباره اعمال می‌کنیم
    # --------------------------------------------------------

    for row in range(
        4,
        worksheet.max_row + 1
    ):

        risk_cell = worksheet.cell(
            row=row,
            column=7
        )

        if risk_cell.value:

            label = str(
                risk_cell.value
            )

            apply_risk_style(
                risk_cell,
                label
            )

            risk_cell.border = styles[
                "border"
            ]

    auto_fit_columns(
        worksheet,
        minimum=12,
        maximum=40
    )


# ============================================================
# RISK CLASS SHEET
# ============================================================

def build_risk_class_sheet(
    workbook,
    values,
    styles
):

    worksheet = workbook.create_sheet(
        "طبقات_خطر"
    )

    add_sheet_title(
        worksheet,
        "توزیع مکانی طبقات خطر FLI",
        styles,
        5
    )

    worksheet.append([])

    worksheet.append([
        "طبقه خطر",
        "حداقل FLI",
        "حداکثر FLI",
        "تعداد سلول",
        "درصد از محدوده معتبر",
    ])

    style_header_row(
        worksheet,
        3,
        styles
    )

    rows = calculate_class_statistics(
        values
    )

    for item in rows:

        worksheet.append([
            item["label"],
            item["min"],
            item["max"],
            item["count"],
            item["percent"],
        ])

        row = worksheet.max_row

        cell = worksheet.cell(
            row=row,
            column=1
        )

        apply_risk_style(
            cell,
            item["label"]
        )

        percent_cell = worksheet.cell(
            row=row,
            column=5
        )

        percent_cell.number_format = "0.00"

    apply_sheet_format(
        worksheet,
        styles
    )

    # Header
    style_header_row(
        worksheet,
        3,
        styles
    )

    # Risk colors
    for row in range(
        4,
        worksheet.max_row + 1
    ):

        cell = worksheet.cell(
            row=row,
            column=1
        )

        if cell.value:

            apply_risk_style(
                cell,
                str(cell.value)
            )

            cell.border = styles[
                "border"
            ]

    auto_fit_columns(
        worksheet,
        minimum=14,
        maximum=35
    )


# ============================================================
# TECHNICAL SHEET
# ============================================================

def build_technical_sheet(
    workbook,
    metadata,
    input_path,
    protected_path,
    hunting_path,
    run_date,
    styles
):

    worksheet = workbook.create_sheet(
        "اطلاعات_فنی"
    )

    add_sheet_title(
        worksheet,
        "اطلاعات فنی و منابع داده",
        styles,
        3
    )

    worksheet.append([])

    worksheet.append([
        "پارامتر",
        "مقدار",
        "توضیح",
    ])

    style_header_row(
        worksheet,
        3,
        styles
    )

    technical_rows = [
        (
            "تاریخ پیش‌بینی",
            run_date or "نامشخص",
            "تاریخ FLI مورد استفاده در گزارش",
        ),
        (
            "Raster ورودی FLI",
            str(input_path),
            "Raster نهایی شاخص خطر حریق",
        ),
        (
            "مرجع مناطق چهارگانه",
            str(protected_path),
            "protected_areas.geojson",
        ),
        (
            "مرجع مناطق شکار ممنوع",
            str(hunting_path),
            "hunting_banned.geojson",
        ),
        (
            "CRS",
            metadata["crs"],
            "سیستم مختصات Raster",
        ),
        (
            "عرض Raster",
            metadata["width"],
            "تعداد ستون‌ها",
        ),
        (
            "ارتفاع Raster",
            metadata["height"],
            "تعداد ردیف‌ها",
        ),
        (
            "تفکیک X",
            metadata["resolution_x"],
            "تفکیک مکانی Raster",
        ),
        (
            "تفکیک Y",
            metadata["resolution_y"],
            "تفکیک مکانی Raster",
        ),
    ]

    for row_data in technical_rows:

        worksheet.append(
            list(row_data)
        )

    apply_sheet_format(
        worksheet,
        styles
    )

    style_header_row(
        worksheet,
        3,
        styles
    )

    auto_fit_columns(
        worksheet,
        minimum=15,
        maximum=60
    )


# ============================================================
# VALIDATE REGION MAPPING
# ============================================================

def validate_region_mapping(
    protected_path: Path,
    hunting_path: Path
):

    protected_name = (
        protected_path.name.lower()
    )

    hunting_name = (
        hunting_path.name.lower()
    )

    if "protected" not in protected_name:

        raise ValueError(
            "Protected areas input does not appear "
            "to be protected_areas.geojson: "
            f"{protected_path}"
        )

    if "hunting" not in hunting_name:

        raise ValueError(
            "Hunting input does not appear "
            "to be hunting_banned.geojson: "
            f"{hunting_path}"
        )


# ============================================================
# MAIN
# ============================================================

def main():

    args = parse_args()

    print("")
    print("=" * 70)
    print("FIRIS EXCEL REPORT")
    print("=" * 70)
    print("")

    # --------------------------------------------------------
    # Validate files
    # --------------------------------------------------------

    require_file(
        args.input,
        "FLI raster"
    )

    require_file(
        args.protected,
        "Protected areas GeoJSON"
    )

    require_file(
        args.hunting,
        "Hunting banned GeoJSON"
    )

    validate_region_mapping(
        args.protected,
        args.hunting
    )

    # --------------------------------------------------------
    # Load FLI
    # --------------------------------------------------------

    print(
        "Reading FLI raster..."
    )

    (
        fli_data,
        fli_stats,
        metadata
    ) = load_fli(
        args.input
    )

    # --------------------------------------------------------
    # Load protected areas
    # --------------------------------------------------------

    print(
        "Reading protected areas..."
    )

    protected_geojson = load_geojson(
        args.protected
    )

    # --------------------------------------------------------
    # Load hunting banned areas
    # --------------------------------------------------------

    print(
        "Reading hunting banned areas..."
    )

    hunting_geojson = load_geojson(
        args.hunting
    )

    # --------------------------------------------------------
    # Check if both source files are identical
    # --------------------------------------------------------

    check_geojson_sources(
        protected_geojson,
        hunting_geojson,
        args.protected,
        args.hunting
    )

    # --------------------------------------------------------
    # Province risk
    # --------------------------------------------------------

    province_risk = risk_class(
        fli_stats["mean"]
    )

    print("")
    print(
        "Province FLI statistics:"
    )

    print(
        f'  Min:  {fli_stats["min"]:.4f}'
    )

    print(
        f'  Max:  {fli_stats["max"]:.4f}'
    )

    print(
        f'  Mean: {fli_stats["mean"]:.4f}'
    )

    print(
        f"  Risk: {province_risk}"
    )

    print("")

    # --------------------------------------------------------
    # Create workbook
    # --------------------------------------------------------

    workbook = Workbook()

    default_sheet = workbook.active

    if default_sheet is not None:

        workbook.remove(
            default_sheet
        )

    styles = create_styles()

    # ========================================================
    # 1. SUMMARY
    # ========================================================

    build_summary_sheet(
        workbook=workbook,
        fli_stats=fli_stats,
        metadata=metadata,
        run_date=args.run_date,
        styles=styles,
    )

    # ========================================================
    # Open raster for regional statistics
    # ========================================================

    with rasterio.open(
        args.input
    ) as src:

        # ====================================================
        # 2. FOUR PROTECTED AREAS
        # ====================================================

        build_region_sheet(
            workbook=workbook,
            sheet_name="مناطق_چهارگانه",
            region_type="مناطق چهارگانه",
            geojson=protected_geojson,
            src=src,
            styles=styles,
        )

        # ====================================================
        # 3. HUNTING BANNED
        # ====================================================

        build_region_sheet(
            workbook=workbook,
            sheet_name="مناطق_شکار_ممنوع",
            region_type="مناطق شکار ممنوع",
            geojson=hunting_geojson,
            src=src,
            styles=styles,
        )

    # ========================================================
    # 4. RISK CLASSES
    # ========================================================

    build_risk_class_sheet(
        workbook=workbook,
        values=fli_data,
        styles=styles,
    )

    # ========================================================
    # 5. TECHNICAL INFORMATION
    # ========================================================

    build_technical_sheet(
        workbook=workbook,
        metadata=metadata,
        input_path=args.input,
        protected_path=args.protected,
        hunting_path=args.hunting,
        run_date=args.run_date,
        styles=styles,
    )

    # ========================================================
    # Sheet order
    # ========================================================

    desired_order = [
        "خلاصه",
        "مناطق_چهارگانه",
        "مناطق_شکار_ممنوع",
        "طبقات_خطر",
        "اطلاعات_فنی",
    ]

    workbook._sheets = [
        workbook[name]
        for name in desired_order
        if name in workbook.sheetnames
    ]

    # ========================================================
    # Final styling pass
    #
    # این مرحله عمداً در آخر انجام می‌شود تا رنگ‌های خطر
    # تحت تأثیر قالب‌بندی عمومی قرار نگیرند.
    # ========================================================

    summary_sheet = workbook["خلاصه"]

    # پیدا کردن طبقه خطر استان
    for row in range(
        1,
        summary_sheet.max_row + 1
    ):

        title_cell = summary_sheet.cell(
            row=row,
            column=1
        )

        if title_cell.value == "طبقه خطر استان":

            risk_cell = summary_sheet.cell(
                row=row,
                column=2
            )

            apply_risk_style(
                risk_cell,
                province_risk
            )

            risk_cell.border = styles[
                "border"
            ]

            break

    # ========================================================
    # Save
    # ========================================================

    args.output.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook.save(
        args.output
    )

    # ========================================================
    # Verification
    # ========================================================

    print("")
    print("=" * 70)
    print("VERIFYING EXCEL")
    print("=" * 70)
    print("")

    print(
        "Output:",
        args.output
    )

    print("")
    print(
        "Sheets:"
    )

    for sheet in workbook.sheetnames:

        print(
            f"  ✓ {sheet}"
        )

    print("")
    print(
        "Region mapping:"
    )

    print(
        "  protected_areas.geojson"
        " -> مناطق چهارگانه"
    )

    print(
        "  hunting_banned.geojson"
        " -> مناطق شکار ممنوع"
    )

    print("")
    print(
        f"Province risk: {province_risk}"
    )

    print("")
    print(
        "FIRIS EXCEL REPORT CREATED SUCCESSFULLY"
    )

    print("=" * 70)
    print("")


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":

    main()
