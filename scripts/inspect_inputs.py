from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import rasterio
except ImportError:
    rasterio = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = PROJECT_ROOT / "data" / "raw"
REPORT_FILE = PROJECT_ROOT / "data" / "input_inventory.json"

RASTER_EXTENSIONS = {".tif", ".tiff"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
TEXT_EXTENSIONS = {".json", ".csv", ".txt", ".xml"}
FWI_FILE_PREFIX = "fwi_ecmwf_fars_"


def json_safe(value: Any) -> Any:
    """Convert values that may not be JSON serializable into safe values."""
    if value is None:
        return None

    if isinstance(value, (str, int, float, bool)):
        return value

    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, TypeError):
            pass

    if isinstance(value, (list, tuple)):
        return [json_safe(item) for item in value]

    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}

    return str(value)


def relative_path(file_path: Path) -> str:
    return file_path.relative_to(PROJECT_ROOT).as_posix()


def inspect_raster(file_path: Path) -> dict[str, Any]:
    result: dict[str, Any] = {
        "kind": "raster",
        "inspection_status": "ok",
    }

    if rasterio is None:
        result["inspection_status"] = "dependency_missing"
        result["message"] = (
            "rasterio is not installed. Raster metadata and statistics were not inspected."
        )
        return result

    try:
        with rasterio.open(file_path) as dataset:
            result.update(
                {
                    "driver": dataset.driver,
                    "width": dataset.width,
                    "height": dataset.height,
                    "band_count": dataset.count,
                    "dtypes": list(dataset.dtypes),
                    "crs": str(dataset.crs) if dataset.crs else None,
                    "bounds": {
                        "left": dataset.bounds.left,
                        "bottom": dataset.bounds.bottom,
                        "right": dataset.bounds.right,
                        "top": dataset.bounds.top,
                    },
                    "transform": list(dataset.transform),
                    "nodata": dataset.nodata,
                    "resolution": {
                        "x": dataset.res[0],
                        "y": dataset.res[1],
                    },
                }
            )

            bands: list[dict[str, Any]] = []

            for band_index in range(1, dataset.count + 1):
                band_info: dict[str, Any] = {"band": band_index}

                try:
                    values = dataset.read(band_index, masked=True)

                    band_info["valid_pixel_count"] = int(values.count())
                    band_info["nodata_pixel_count"] = int(values.size - values.count())

                    if values.count() > 0:
                        band_info["minimum"] = json_safe(values.min())
                        band_info["maximum"] = json_safe(values.max())
                        band_info["mean"] = json_safe(values.mean())
                    else:
                        band_info["minimum"] = None
                        band_info["maximum"] = None
                        band_info["mean"] = None

                except Exception as error:
                    band_info["statistics_error"] = str(error)

                bands.append(band_info)

            result["bands"] = bands

    except Exception as error:
        result["inspection_status"] = "error"
        result["message"] = str(error)

    return result


def inspect_excel(file_path: Path) -> dict[str, Any]:
    result: dict[str, Any] = {
        "kind": "excel",
        "inspection_status": "ok",
    }

    if load_workbook is None:
        result["inspection_status"] = "dependency_missing"
        result["message"] = (
            "openpyxl is not installed. Excel workbook structure was not inspected."
        )
        return result

    try:
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=False,
        )

        sheets: list[dict[str, Any]] = []

        for worksheet in workbook.worksheets:
            headers: list[Any] = []

            for row in worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ):
                headers = [json_safe(value) for value in row]
                break

            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "first_row_values": headers,
                }
            )

        workbook.close()
        result["sheets"] = sheets

    except Exception as error:
        result["inspection_status"] = "error"
        result["message"] = str(error)

    return result


def inspect_text_file(file_path: Path) -> dict[str, Any]:
    result: dict[str, Any] = {
        "kind": "text_or_structured",
        "inspection_status": "ok",
    }

    try:
        result["preview"] = file_path.read_text(
            encoding="utf-8",
            errors="replace",
        )[:1000]
    except Exception as error:
        result["inspection_status"] = "error"
        result["message"] = str(error)

    return result


def inspect_file(file_path: Path) -> dict[str, Any]:
    suffix = file_path.suffix.lower()

    record: dict[str, Any] = {
        "path": relative_path(file_path),
        "name": file_path.name,
        "extension": suffix,
        "size_bytes": file_path.stat().st_size,
    }

    if suffix in RASTER_EXTENSIONS:
        record.update(inspect_raster(file_path))
    elif suffix in EXCEL_EXTENSIONS:
        record.update(inspect_excel(file_path))
    elif suffix in TEXT_EXTENSIONS:
        record.update(inspect_text_file(file_path))
    else:
        record.update(
            {
                "kind": "unknown",
                "inspection_status": "not_inspected",
                "message": (
                    "No inspector is configured for this file extension. "
                    "The file was listed but not opened."
                ),
            }
        )

    return record


def main() -> int:
    if not RAW_DIR.exists():
        print(f"Raw data directory does not exist: {RAW_DIR}")
        return 1

    files = sorted(
        path
        for path in RAW_DIR.rglob("*")
        if path.is_file() and path.name != "keep.txt"
    )

    inventory: dict[str, Any] = {
        "project": "FIRIS",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "raw_data_directory": relative_path(RAW_DIR),
        "file_count": len(files),
        "files": [inspect_file(file_path) for file_path in files],
    }

    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    REPORT_FILE.write_text(
        json.dumps(inventory, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"Inspection complete. Files found: {len(files)}")
    print(f"Report written: {REPORT_FILE}")

    if not files:
        print(
            "No raw input files were found. "
            "Add fars_fuel.tif, the fuel Excel file, dem_fars.tif, "
            "or download FWI before running the build step."
        )

    return 0


if __name__ == "__main__":
    sys.exit(main())
